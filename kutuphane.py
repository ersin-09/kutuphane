# -*- coding: utf-8 -*-
import os, sys, sqlite3, datetime, re, shutil, random, string
from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QTabWidget, QVBoxLayout, QHBoxLayout,
    QFormLayout, QLineEdit, QPushButton, QMessageBox, QTableWidget, QTableWidgetItem,
    QLabel, QSpinBox, QDateEdit, QComboBox, QFileDialog, QGroupBox, QListWidget, QDialog
)
from PyQt5.QtCore import QDate, Qt
from openpyxl import Workbook, load_workbook
from reportlab.pdfgen import canvas as pdfcanvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
import hashlib

PRIMARY_BUTTON_STYLE = """
QPushButton {
    background-color: #2563EB;
    color: white;
    font-weight: 600;
    border-radius: 6px;
    padding: 8px 18px;
}
QPushButton:hover {
    background-color: #1D4ED8;
}
QPushButton:pressed {
    background-color: #1E40AF;
}
QPushButton:disabled {
    background-color: #93C5FD;
    color: white;
}
"""

SECONDARY_BUTTON_STYLE = """
QPushButton {
    background-color: #64748B;
    color: white;
    font-weight: 500;
    border-radius: 6px;
    padding: 6px 14px;
}
QPushButton:hover {
    background-color: #475569;
}
QPushButton:pressed {
    background-color: #334155;
}
QPushButton:disabled {
    background-color: #CBD5F5;
    color: white;
}
"""

DANGER_BUTTON_STYLE = """
QPushButton {
    background-color: #DC2626;
    color: white;
    font-weight: 600;
    border-radius: 6px;
    padding: 8px 18px;
}
QPushButton:hover {
    background-color: #B91C1C;
}
QPushButton:pressed {
    background-color: #991B1B;
}
QPushButton:disabled {
    background-color: #FCA5A5;
    color: white;
}
"""


def _apply_button_style(button: QPushButton, stylesheet: str, min_height: int) -> None:
    button.setStyleSheet(stylesheet)
    button.setMinimumHeight(min_height)
    button.setCursor(QtGui.QCursor(Qt.PointingHandCursor))


def style_primary(button: QPushButton, min_height: int = 40) -> None:
    _apply_button_style(button, PRIMARY_BUTTON_STYLE, min_height)


def style_secondary(button: QPushButton, min_height: int = 32) -> None:
    _apply_button_style(button, SECONDARY_BUTTON_STYLE, min_height)


def style_danger(button: QPushButton, min_height: int = 40) -> None:
    _apply_button_style(button, DANGER_BUTTON_STYLE, min_height)

# -------------------------------------------------------------------------
# Uygulamanın ana dizinini kesin olarak belirleme (.py ve .exe için)
# -------------------------------------------------------------------------
if getattr(sys, 'frozen', False):
    # Eğer uygulama bir .exe içindeyse, ana dizin sys.argv[0] ile alınır.
    APP_DIR = os.path.dirname(sys.argv[0])
else:
    # Eğer bir .py dosyasından çalışıyorsa, ana dizin dosyanın kendisinin bulunduğu yerdir.
    APP_DIR = os.path.dirname(os.path.abspath(__file__))

DB_DIR = os.path.join(APP_DIR, "db")
DB_PATH = os.path.join(DB_DIR, "kutuphane.db")
BACKUP_DIR = os.path.join(APP_DIR, "yedek")
EXPORT_DIR = os.path.join(APP_DIR, "disa_aktar")
IMPORT_DIR = os.path.join(APP_DIR, "ice_aktar")
REPORT_DIR = os.path.join(APP_DIR, "raporlar")
# -------------------------------------------------------------------------

# Türkçe karakter desteği için font tanımlaması
try:
    FONT_PATH = "C:\\Windows\\Fonts\\Arial.ttf" # Windows için varsayılan yol
    if not os.path.exists(FONT_PATH):
        # MacOS için varsayılan yol
        FONT_PATH = "/Library/Fonts/Arial.ttf"
        if not os.path.exists(FONT_PATH):
            FONT_PATH = None
            print("Uyarı: Arial.ttf fontu bulunamadı. PDF'lerde Türkçe karakter sorunu yaşanabilir.")
            print("Lütfen FONT_PATH değişkenini sisteminizdeki bir fontun yoluyla güncelleyin.")
    if FONT_PATH:
        pdfmetrics.registerFont(TTFont('Arial', FONT_PATH))
        pdfmetrics.registerFont(TTFont('Arial-Bold', FONT_PATH)) # Aynı fontu kalın font için de kullanabiliriz
        pdfmetrics.registerFontFamily('Arial', normal='Arial', bold='Arial-Bold', italic='Arial', boldItalic='Arial')
        DEFAULT_FONT = 'Arial'
    else:
        DEFAULT_FONT = 'Helvetica' # Varsayılan olarak Helvetica'ya dön
except Exception as e:
    DEFAULT_FONT = 'Helvetica'
    print(f"PDF font hatası: {e}. Helvetica'ya geçiliyor.")


def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def ensure_dirs():
    for d in [DB_DIR, BACKUP_DIR, EXPORT_DIR, IMPORT_DIR, REPORT_DIR]:
        os.makedirs(d, exist_ok=True)

def auto_backup():
    """
    Uygulama başladığında mevcut veritabanını yedekler ve
    yedek dosya sayısını belirlenen limitte tutar.
    """
    ensure_dirs()
    if not os.path.exists(DB_PATH):
        print("Veritabanı dosyası bulunamadı, yedekleme atlandı.")
        return

    # Yedekleme işlemini yap
    try:
        ts = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        dst = os.path.join(BACKUP_DIR, f"{ts}_kutuphane.db")
        shutil.copy2(DB_PATH, dst)
        print("Otomatik yedekleme tamamlandı:", dst)
    except Exception as e:
        print("Otomatik yedekleme hatası:", e)
        return

    # --- Yedekleme temizleme politikası ---
    try:
        max_backups = 20
        
        backup_files = [f for f in os.listdir(BACKUP_DIR) if f.endswith("_kutuphane.db")]
        
        backup_files.sort(key=lambda f: os.path.getmtime(os.path.join(BACKUP_DIR, f)))
        
        if len(backup_files) > max_backups:
            files_to_delete = backup_files[:-max_backups]
            for file_name in files_to_delete:
                file_path = os.path.join(BACKUP_DIR, file_name)
                os.remove(file_path)
                print(f"Eski yedek silindi: {file_name}")

    except Exception as e:
        print("Eski yedekleri silme hatası:", e)

def normalize(s: str) -> str:
    if s is None: return ""
    t = str(s).strip().lower()
    REP = {"ı":"i","İ":"i","ş":"s","Ş":"s","ğ":"g","Ğ":"g","ü":"u","Ü":"u","ö":"o","Ö":"o","ç":"c","Ç":"c"}
    for k,v in REP.items(): t = t.replace(k,v)
    t = re.sub(r"\s+", " ", t)
    return t

def db_conn():
    ensure_dirs()
    conn = sqlite3.connect(DB_PATH)
    conn.create_function("normalize", 1, normalize)
    return conn

def init_db():
    ensure_dirs()
    conn = db_conn()
    c = conn.cursor()
    # books table
    c.execute("""CREATE TABLE IF NOT EXISTS books(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        barcode TEXT UNIQUE,
        title TEXT,
        author TEXT,
        publisher TEXT,
        year INTEGER,
        pages INTEGER,
        category TEXT,
        demirbas TEXT,
        raf TEXT,
        dolap TEXT,
        adet INTEGER DEFAULT 1,
        note TEXT
    )""")
    # members table
    c.execute("""CREATE TABLE IF NOT EXISTS members(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT,
        surname TEXT,
        class TEXT,
        branch TEXT,
        no TEXT UNIQUE,
        gender TEXT,
        phone TEXT,
        register_date TEXT
    )""")
    # loans table
    c.execute("""CREATE TABLE IF NOT EXISTS loans(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        book_id INTEGER,
        member_id INTEGER,
        loan_date TEXT,
        due_date TEXT,
        return_date TEXT,
        FOREIGN KEY(book_id) REFERENCES books(id),
        FOREIGN KEY(member_id) REFERENCES members(id)
    )""")
    # Yeni users tablosu
    c.execute("""CREATE TABLE IF NOT EXISTS users(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE,
        password TEXT,
        role TEXT DEFAULT 'staff'
    )""")
    
    # Yeni settings tablosu
    c.execute("""CREATE TABLE IF NOT EXISTS settings(
        key TEXT PRIMARY KEY,
        value TEXT
    )""")
    
    # Varsayılan kullanıcıyı ekle
    c.execute("SELECT COUNT(*) FROM users WHERE username = 'Admin'")
    if c.fetchone()[0] == 0:
        hashed_password = hash_password("12345")
        c.execute("INSERT INTO users (username, password, role) VALUES (?, ?, ?)", ('Admin', hashed_password, 'admin'))
    
    # Kullanıcı tablosunda role sütunu yoksa ekle (eski veritabanı için)
    try:
        c.execute("SELECT role FROM users LIMIT 1")
    except sqlite3.OperationalError:
        c.execute("ALTER TABLE users ADD COLUMN role TEXT DEFAULT 'staff'")
        c.execute("UPDATE users SET role = 'admin' WHERE username = 'Admin'")

    # Ödünç alma sınırı ayarını ekle (varsayılan 3)
    c.execute("SELECT COUNT(*) FROM settings WHERE key = 'loan_limit'")
    if c.fetchone()[0] == 0:
        c.execute("INSERT INTO settings (key, value) VALUES (?, ?)", ('loan_limit', '3'))

    # YENİ EKLEME: Ödünç verme süresi ayarını ekle (varsayılan 15)
    c.execute("SELECT COUNT(*) FROM settings WHERE key = 'default_loan_days'")
    if c.fetchone()[0] == 0:
        c.execute("INSERT INTO settings (key, value) VALUES (?, ?)", ('default_loan_days', '15'))
    
    conn.commit()
    conn.close()

TURKISH_HEADER_MAP = {
    "barkod":"barcode",
    "kitapadi":"title","kitap adi":"title","kitap adı":"title",
    "yazar":"author",
    "yayınevi":"publisher","yayinevi":"publisher","yayınevi̇":"publisher",
    "basimyili":"year","basim yili":"year","basim yılı":"year","basım yılı":"year",
    "sayfasayisi":"pages","sayfa sayisi":"pages","sayfa sayısı":"pages",
    "turu":"category","türü":"category","kategori":"category",
    "demirbas":"demirbas","demirbasnumarasi":"demirbas","demirbaş numarası":"demirbas",
    "raf":"raf","rafnumarasi":"raf","raf numarası":"raf",
    "dolap":"dolap",
    "adet":"adet","adedi":"adet",
    "aciklama":"note","açiklama":"note","açıklama":"note",
}

def gen_barcode(n=8):
    return "".join(random.choice(string.digits) for _ in range(n))

def excel_val_to_iso(v):
    if v is None or v=="": return ""
    if isinstance(v, datetime.datetime): return v.date().isoformat()
    if isinstance(v, datetime.date): return v.isoformat()
    s = str(v).strip()
    for sep in [".","/","-"]:
        parts = s.split(sep)
        if len(parts)==3:
            try:
                if len(parts[0])==4:
                    y,m,d = int(parts[0]), int(parts[1]), int(parts[2])
                else:
                    d,m,y = int(parts[0]), int(parts[1]), int(parts[2])
                return datetime.date(y,m,d).isoformat()
            except: pass
    return s

# YENİ EKLEME: Ayar değerlerini okumak için yardımcı fonksiyonlar
def get_setting(key: str, default: str) -> str:
    with db_conn() as conn:
        c = conn.cursor()
        c.execute("SELECT value FROM settings WHERE key = ?", (key,))
        result = c.fetchone()
        if result:
            return result[0]
        return default

def get_loan_limit() -> int:
    try:
        return int(get_setting('loan_limit', '3'))
    except (ValueError, IndexError):
        return 3

def get_default_loan_days() -> int:
    try:
        return int(get_setting('default_loan_days', '15'))
    except (ValueError, IndexError):
        return 15

# -------------------- Books Tab --------------------
class BooksTab(QWidget):
    def __init__(self, user_role, parent=None):
        super().__init__(parent)
        self.user_role = user_role
        self.current_sort_column = -1
        self.sort_order = QtCore.Qt.AscendingOrder
        self.build_ui()
        if self.user_role == 'staff':
            self.btnDelete.setDisabled(True)
            self.btnDeleteAll.hide()

    def build_ui(self):
        main = QVBoxLayout(self)

        splitter = QtWidgets.QSplitter(QtCore.Qt.Vertical)

        top_panel = QWidget()
        top_layout = QVBoxLayout(top_panel)
        
        form_layout = QHBoxLayout()

        form_box1 = QGroupBox("Temel Kitap Bilgileri")
        form1 = QFormLayout(form_box1)
        self.edBarcode = QLineEdit(); self.edBarcode.setReadOnly(True)
        self.edTitle = QLineEdit()
        self.edAuthor = QLineEdit()
        self.edPublisher = QLineEdit()
        self.spYear = QSpinBox(); self.spYear.setRange(0, 2100); self.spYear.setValue(2020)
        self.spPages = QSpinBox(); self.spPages.setRange(0, 10000)
        
        form1.addRow("Barkod:", self.edBarcode)
        form1.addRow("Kitap Adı:", self.edTitle)
        form1.addRow("Yazar:", self.edAuthor)
        form1.addRow("Yayınevi:", self.edPublisher)
        form1.addRow("Basım Yılı:", self.spYear)
        form1.addRow("Sayfa Sayısı:", self.spPages)

        form_box2 = QGroupBox("Ek Bilgiler")
        form2 = QFormLayout(form_box2)
        self.edCategory = QLineEdit()
        self.edDemirbas = QLineEdit(); self.edDemirbas.setReadOnly(True)
        self.edRaf = QLineEdit()
        self.edDolap = QLineEdit()
        self.spAdet = QSpinBox(); self.spAdet.setRange(0, 1000); self.spAdet.setValue(1)
        self.edNote = QLineEdit()
        
        form2.addRow("Türü:", self.edCategory)
        form2.addRow("Demirbaş No:", self.edDemirbas)
        form2.addRow("Raf:", self.edRaf)
        form2.addRow("Dolap:", self.edDolap)
        form2.addRow("Adet:", self.spAdet)
        form2.addRow("Açıklama:", self.edNote)

        form_layout.addWidget(form_box1)
        form_layout.addWidget(form_box2)
        
        top_layout.addLayout(form_layout)
        
        btn_row = QHBoxLayout()
        self.btnAdd = QPushButton("Ekle"); style_primary(self.btnAdd)
        self.btnUpdate = QPushButton("Güncelle"); style_primary(self.btnUpdate)
        self.btnDelete = QPushButton("Seçiliyi Sil"); style_danger(self.btnDelete)
        self.btnDeleteAll = QPushButton("Tümünü Sil"); style_danger(self.btnDeleteAll)
        self.btnClear = QPushButton("Temizle"); style_secondary(self.btnClear)
        btn_row.addWidget(self.btnAdd); btn_row.addWidget(self.btnUpdate)
        btn_row.addWidget(self.btnDelete); btn_row.addWidget(self.btnDeleteAll)
        btn_row.addStretch(1)
        btn_row.addWidget(self.btnClear)
        top_layout.addLayout(btn_row)

        search_box = QGroupBox("Arama ve Dosya İşlemleri")
        sh = QHBoxLayout(search_box)
        self.edSearch = QLineEdit(); self.edSearch.setPlaceholderText("Başlık, yazar veya barkoda göre ara")
        self.btnSearch = QPushButton("Ara"); style_secondary(self.btnSearch)
        
        filter_layout = QHBoxLayout()
        self.cbFilterCategory = QComboBox(); self.cbFilterCategory.addItem("Tüm Kategoriler")
        self.cbFilterAuthor = QComboBox(); self.cbFilterAuthor.addItem("Tüm Yazarlar")
        self.btnFilter = QPushButton("Filtrele"); style_secondary(self.btnFilter)

        filter_layout.addWidget(QLabel("Kategori:")); filter_layout.addWidget(self.cbFilterCategory)
        filter_layout.addWidget(QLabel("Yazar:")); filter_layout.addWidget(self.cbFilterAuthor)
        filter_layout.addWidget(self.btnFilter)
        
        sh.addWidget(self.edSearch); sh.addWidget(self.btnSearch)
        sh.addLayout(filter_layout)
        sh.addStretch(1)
        
        file_ops_layout = QHBoxLayout()
        self.btnExport = QPushButton("Excel'e Aktar"); style_secondary(self.btnExport)
        self.btnImport = QPushButton("Excel'den Al"); style_secondary(self.btnImport)
        file_ops_layout.addWidget(self.btnExport); file_ops_layout.addWidget(self.btnImport)
        sh.addLayout(file_ops_layout)
        
        top_layout.addWidget(search_box)
        
        bottom_panel = QWidget()
        bottom_layout = QVBoxLayout(bottom_panel)
        self.tbl = QTableWidget(0, 8)
        self.tbl.setHorizontalHeaderLabels(["ID","Barkod","Ad","Yazar","Yayınevi","Yıl","Adet","Tür"])
        self.tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl.setSelectionMode(QTableWidget.ExtendedSelection)
        self.tbl.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        bottom_layout.addWidget(self.tbl)

        splitter.addWidget(top_panel)
        splitter.addWidget(bottom_panel)
        main.addWidget(splitter)

        self.btnAdd.clicked.connect(self.on_add)
        self.btnUpdate.clicked.connect(self.on_update)
        self.btnDelete.clicked.connect(self.on_delete)
        self.btnDeleteAll.clicked.connect(self.on_delete_all)
        self.btnClear.clicked.connect(self.clear_form)
        self.btnSearch.clicked.connect(self.refresh)
        self.btnFilter.clicked.connect(self.refresh)
        self.btnExport.clicked.connect(self.export_excel)
        self.btnImport.clicked.connect(self.import_excel)
        self.tbl.itemSelectionChanged.connect(self.fill_form)
        self.tbl.horizontalHeader().sectionClicked.connect(self.sort_table)
        
        self.load_filters()
        self.refresh()

    def sort_table(self, column):
        if column == self.current_sort_column:
            self.sort_order = QtCore.Qt.AscendingOrder if self.sort_order == QtCore.Qt.DescendingOrder else QtCore.Qt.DescendingOrder
        else:
            self.current_sort_column = column
            self.sort_order = QtCore.Qt.AscendingOrder
        self.tbl.sortItems(self.current_sort_column, self.sort_order)

    def load_filters(self):
        with db_conn() as conn:
            c = conn.cursor()
            categories = c.execute("SELECT DISTINCT category FROM books WHERE category IS NOT NULL ORDER BY category").fetchall()
            authors = c.execute("SELECT DISTINCT author FROM books WHERE author IS NOT NULL ORDER BY author").fetchall()
            self.cbFilterCategory.addItems([c[0] for c in categories])
            self.cbFilterAuthor.addItems([a[0] for a in authors])

    def current_row_id(self):
        rows = self.tbl.selectionModel().selectedRows()
        if not rows: return None
        r = rows[0].row()
        return int(self.tbl.item(r,0).text())

    def fill_form(self):
        rid = self.current_row_id()
        if rid is None:
            self.clear_form()
            return
        with db_conn() as conn:
            c = conn.cursor()
            c.execute("SELECT barcode,title,author,publisher,year,pages,category,demirbas,raf,dolap,adet,note FROM books WHERE id=?", (rid,))
            row = c.fetchone()
            if not row: return
            (barcode,title,author,publisher,year,pages,category,demirbas,raf,dolap,adet,note) = row
            self.edBarcode.setText("" if barcode is None else str(barcode))
            self.edTitle.setText(title or "")
            self.edAuthor.setText(author or "")
            self.edPublisher.setText(publisher or "")
            self.spYear.setValue(year or 0)
            self.spPages.setValue(pages or 0)
            self.edCategory.setText(category or "")
            self.edDemirbas.setText(demirbas or "")
            self.edRaf.setText(raf or "")
            self.edDolap.setText(dolap or "")
            self.spAdet.setValue(adet or 0)
            self.edNote.setText(note or "")
            self.btnUpdate.setDisabled(False)
            self.btnAdd.setDisabled(True)

    def clear_form(self):
        for w in [self.edBarcode,self.edTitle,self.edAuthor,self.edPublisher,self.edCategory,self.edDemirbas,self.edRaf,self.edDolap,self.edNote]:
            w.clear()
        self.spYear.setValue(2020)
        self.spPages.setValue(0)
        self.spAdet.setValue(1)
        self.tbl.clearSelection()
        self.btnUpdate.setDisabled(True)
        self.btnAdd.setDisabled(False)

    def on_add(self):
        title = self.edTitle.text().strip()
        dolap = self.edDolap.text().strip()
        raf = self.edRaf.text().strip()
        
        if not title:
            QMessageBox.warning(self, "Uyarı", "Kitap adı zorunludur.")
            return

        if not (dolap and raf):
            QMessageBox.warning(self, "Uyarı", "Otomatik barkod ve demirbaş numarası oluşturulması için 'Dolap' ve 'Raf' bilgileri zorunludur.")
            return

        barcode = f"{dolap}-{raf}"
        demirbas = f"{dolap}-{raf}"

        data = (
            barcode, title, self.edAuthor.text().strip() or None, self.edPublisher.text().strip() or None,
            int(self.spYear.value()) or None, int(self.spPages.value()) or None,
            self.edCategory.text().strip() or None, demirbas or None,
            raf or None, dolap or None,
            int(self.spAdet.value()), self.edNote.text().strip() or None
        )
        try:
            with db_conn() as conn:
                c = conn.cursor()
                c.execute("""INSERT INTO books(barcode,title,author,publisher,year,pages,category,demirbas,raf,dolap,adet,note)
                             VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""", data)
                conn.commit()
            self.refresh(); self.clear_form()
            QMessageBox.information(self, "Başarılı", "Kitap başarıyla eklendi.")
        except sqlite3.IntegrityError:
            QMessageBox.warning(self, "Hata", "Aynı barkod/demirbaş numarası zaten mevcut. Lütfen dolap ve raf bilgilerini kontrol edin.")

    def on_update(self):
        rid = self.current_row_id()
        if rid is None:
            QMessageBox.information(self, "Bilgi", "Güncellenecek satırı seçin.")
            return
        
        title = self.edTitle.text().strip()
        dolap = self.edDolap.text().strip()
        raf = self.edRaf.text().strip()

        if not title:
            QMessageBox.warning(self, "Uyarı", "Kitap adı zorunludur.")
            return
        if not (dolap and raf):
            QMessageBox.warning(self, "Uyarı", "Otomatik barkod ve demirbaş numarası oluşturulması için 'Dolap' ve 'Raf' bilgileri zorunludur.")
            return

        barcode = f"{dolap}-{raf}"
        demirbas = f"{dolap}-{raf}"

        data = (
            barcode, title, self.edAuthor.text().strip() or None, self.edPublisher.text().strip() or None,
            int(self.spYear.value()) or None, int(self.spPages.value()) or None,
            self.edCategory.text().strip() or None, demirbas or None,
            raf or None, dolap or None,
            int(self.spAdet.value()), self.edNote.text().strip() or None, rid
        )
        try:
            with db_conn() as conn:
                c = conn.cursor()
                c.execute("""UPDATE books SET barcode=?,title=?,author=?,publisher=?,year=?,pages=?,
                             category=?,demirbas=?,raf=?,dolap=?,adet=?,note=? WHERE id=?""", data)
                conn.commit()
            self.refresh()
            QMessageBox.information(self, "Başarılı", "Kitap bilgileri başarıyla güncellendi.")
        except sqlite3.IntegrityError:
            QMessageBox.warning(self, "Hata", "Güncelleme için girilen barkod/demirbaş numarası zaten mevcut.")

    def on_delete(self):
        selected_rows = self.tbl.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.information(self, "Bilgi", "Lütfen silmek için bir veya daha fazla kitap seçin.")
            return

        reply = QMessageBox.question(self, "Silme Onayı",
                                     f"Seçili {len(selected_rows)} kitabı silmek istediğinizden emin misiniz?",
                                     QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            ids_to_delete = [int(self.tbl.item(r.row(), 0).text()) for r in selected_rows]
            with db_conn() as conn:
                c = conn.cursor()
                c.execute("DELETE FROM books WHERE id IN ({})".format(','.join('?' for _ in ids_to_delete)), ids_to_delete)
                conn.commit()
            self.refresh()
            self.clear_form()
            QMessageBox.information(self, "Başarılı", f"{len(ids_to_delete)} kitap başarıyla silindi.")

    def on_delete_all(self):
        if self.user_role != 'admin':
            QMessageBox.warning(self, "Yetkisiz İşlem", "Bu işlemi yapmak için yönetici yetkisi gereklidir.")
            return
        reply = QMessageBox.question(self, "Tüm Kitapları Sil",
                                    "**DİKKAT!** Bu işlem tüm kitapları kalıcı olarak silecektir. Emin misiniz?",
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            with db_conn() as conn:
                c = conn.cursor()
                c.execute("DELETE FROM books")
                conn.commit()
            self.refresh()
            self.clear_form()
            QMessageBox.information(self, "Tamamlandı", "Tüm kitaplar başarıyla silindi.")

    def export_excel(self):
        path, _ = QFileDialog.getSaveFileName(self, "Excel'e aktar", os.path.join(EXPORT_DIR, "kitaplar.xlsx"), "Excel (*.xlsx)")
        if not path: return
        headers = ["Barkod","Kitap Adı","Yazar","Yayınevi","Basım Yılı","Sayfa Sayısı","Tür","Demirbaş","Raf","Dolap","Adet","Açıklama"]
        with db_conn() as conn:
            c = conn.cursor()
            c.execute("""SELECT barcode,title,author,publisher,year,pages,category,demirbas,raf,dolap,adet,note FROM books""")
            rows = c.fetchall()
        wb = Workbook(); ws = wb.active; ws.append(headers)
        for r in rows:
            ws.append(list(r))
        wb.save(path)
        QMessageBox.information(self, "Tamam", "Excel dosyası oluşturuldu.")

    def import_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "Excel'den al", IMPORT_DIR, "Excel (*.xlsx)")
        if not path: return
        wb = load_workbook(path); ws = wb.active
        headers = [normalize(str(cell.value)) for cell in ws[1]]
        colmap = {}
        for idx, h in enumerate(headers):
            if h in TURKISH_HEADER_MAP:
                colmap[TURKISH_HEADER_MAP[h]] = idx+1
        needed = ["title"]
        if not all(n in colmap for n in needed):
            QMessageBox.warning(self, "Hata", "Excel başlıkları eksik (en azından 'Kitap Adı' gerekli).")
            return
        count=0
        with db_conn() as conn:
            c = conn.cursor()
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                title = row[colmap["title"]-1]
                if not title: continue
                barcode = row[colmap.get("barcode")-1] if colmap.get("barcode") else None
                if not barcode or str(barcode).strip()=="" : barcode = gen_barcode()
                author = row[colmap.get("author")-1] if colmap.get("author") else None
                publisher = row[colmap.get("publisher")-1] if colmap.get("publisher") else None
                year = row[colmap.get("year")-1] if colmap.get("year") else None
                pages = row[colmap.get("pages")-1] if colmap.get("pages") else None
                category = row[colmap.get("category")-1] if colmap.get("category") else None
                demirbas = row[colmap.get("demirbas")-1] if colmap.get("demirbas") else None
                raf = row[colmap.get("raf")-1] if colmap.get("raf") else None
                dolap = row[colmap.get("dolap")-1] if colmap.get("dolap") else None
                adet = row[colmap.get("adet")-1] if colmap.get("adet") else 1
                note = row[colmap.get("note")-1] if colmap.get("note") else None
                try:
                    c.execute("""INSERT OR IGNORE INTO books(barcode,title,author,publisher,year,pages,category,demirbas,raf,dolap,adet,note)
                                 VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
                             (str(barcode).strip(), str(title).strip(), author, publisher,
                              int(year) if year not in (None,"") else None,
                              int(pages) if pages not in (None,"") else None,
                              category, demirbas, raf, dolap,
                              int(adet) if adet not in (None,"") else 1,
                              note))
                    count += c.rowcount
                except Exception as e:
                    print("Satır", i, "hata:", e)
            conn.commit()
        self.refresh()
        QMessageBox.information(self, "Tamam", f"Excel'den {count} kayıt eklendi (mevcut barkodlar atlandı).")

    def refresh(self):
        q = normalize(self.edSearch.text())
        filter_cat = self.cbFilterCategory.currentText()
        filter_auth = self.cbFilterAuthor.currentText()
        
        query = """SELECT id,barcode,title,author,publisher,year,adet,category FROM books """
        params = []
        where_clauses = []

        if q:
            where_clauses.append("(normalize(title) LIKE ? OR normalize(author) LIKE ? OR barcode LIKE ?)")
            params.extend([f"%{q}%", f"%{q}%", f"%{q}%"])
        if filter_cat != "Tüm Kategoriler":
            where_clauses.append("category = ?")
            params.append(filter_cat)
        if filter_auth != "Tüm Yazarlar":
            where_clauses.append("author = ?")
            params.append(filter_auth)
        
        if where_clauses:
            query += " WHERE " + " AND ".join(where_clauses)
            
        query += " ORDER BY title ASC"
        
        with db_conn() as conn:
            c = conn.cursor()
            c.execute(query, params)
            rows = c.fetchall()
        
        self.tbl.setRowCount(len(rows))
        for r, row in enumerate(rows):
            for cidx, val in enumerate(row):
                item = QTableWidgetItem("" if val is None else str(val))
                if isinstance(val, int): item.setTextAlignment(Qt.AlignCenter)
                self.tbl.setItem(r, cidx, item)
        self.tbl.resizeColumnsToContents()

# -------------------- Members Tab --------------------
class MembersTab(QWidget):
    def __init__(self, user_role, parent=None):
        super().__init__(parent)
        self.user_role = user_role
        self.current_sort_column = -1
        self.sort_order = QtCore.Qt.AscendingOrder
        self.build_ui()
        if self.user_role == 'staff':
            self.btnDelete.setDisabled(True)
            self.btnDeleteAll.hide()

    def build_ui(self):
        main = QVBoxLayout(self)
        
        splitter = QtWidgets.QSplitter(QtCore.Qt.Vertical)

        top_panel = QWidget()
        top_layout = QVBoxLayout(top_panel)
        
        form_layout = QHBoxLayout()
        
        form_box1 = QGroupBox("Kişisel Bilgiler")
        form1 = QFormLayout(form_box1)
        self.edName = QLineEdit()
        self.edSurname = QLineEdit()
        self.edClass = QLineEdit()
        self.edBranch = QLineEdit()
        self.edNo = QLineEdit()
        
        form1.addRow("Ad:", self.edName)
        form1.addRow("Soyad:", self.edSurname)
        form1.addRow("Sınıf:", self.edClass)
        form1.addRow("Şube:", self.edBranch)
        form1.addRow("Numara:", self.edNo)

        form_box2 = QGroupBox("İletişim ve Kayıt Bilgileri")
        form2 = QFormLayout(form_box2)
        self.cbGender = QComboBox(); self.cbGender.addItems(["","Erkek","Kız"])
        self.edPhone = QLineEdit()
        self.deRegister = QDateEdit(); self.deRegister.setCalendarPopup(True); self.deRegister.setDate(QDate.currentDate())

        form2.addRow("Cinsiyet:", self.cbGender)
        form2.addRow("Telefon:", self.edPhone)
        form2.addRow("Kayıt Tarihi:", self.deRegister)

        form_layout.addWidget(form_box1)
        form_layout.addWidget(form_box2)
        top_layout.addLayout(form_layout)

        btn_row = QHBoxLayout()
        self.btnAdd = QPushButton("Ekle"); style_primary(self.btnAdd)
        self.btnUpdate = QPushButton("Güncelle"); style_primary(self.btnUpdate)
        self.btnDelete = QPushButton("Seçiliyi Sil"); style_danger(self.btnDelete)
        self.btnClear = QPushButton("Temizle"); style_secondary(self.btnClear)
        self.btnDeleteAll = QPushButton("Tümünü Sil"); style_danger(self.btnDeleteAll)
        self.btnUpdate.setDisabled(True)
        btn_row.addWidget(self.btnAdd); btn_row.addWidget(self.btnUpdate)
        btn_row.addWidget(self.btnDelete); btn_row.addWidget(self.btnClear)
        btn_row.addStretch(1)
        btn_row.addWidget(self.btnDeleteAll)
        top_layout.addLayout(btn_row)

        search_box = QGroupBox("Arama ve Dosya İşlemleri")
        sh = QHBoxLayout(search_box)
        self.edSearch = QLineEdit(); self.edSearch.setPlaceholderText("Ad, soyad veya numaraya göre ara")
        self.btnSearch = QPushButton("Ara"); style_secondary(self.btnSearch)
        self.btnExport = QPushButton("Excel'e Aktar"); style_secondary(self.btnExport)
        self.btnImport = QPushButton("Excel'den Al"); style_secondary(self.btnImport)
        sh.addWidget(self.edSearch); sh.addWidget(self.btnSearch); sh.addWidget(self.btnExport); sh.addWidget(self.btnImport)
        top_layout.addWidget(search_box)
        
        bottom_panel = QWidget()
        bottom_layout = QVBoxLayout(bottom_panel)
        self.tbl = QTableWidget(0, 8)
        self.tbl.setHorizontalHeaderLabels(["Numara","Ad","Soyad","Sınıf","Şube","Cinsiyet","Telefon","Kayıt Tarihi"])
        self.tbl.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl.setSelectionMode(QTableWidget.ExtendedSelection)
        self.tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        bottom_layout.addWidget(self.tbl)

        splitter.addWidget(top_panel)
        splitter.addWidget(bottom_panel)
        main.addWidget(splitter)

        self.btnAdd.clicked.connect(self.on_add)
        self.btnUpdate.clicked.connect(self.on_update)
        self.btnDelete.clicked.connect(self.on_delete)
        self.btnClear.clicked.connect(self.clear_form)
        self.btnDeleteAll.clicked.connect(self.on_delete_all_members)
        self.btnSearch.clicked.connect(self.refresh)
        self.btnExport.clicked.connect(self.export_excel)
        self.btnImport.clicked.connect(self.import_excel)
        self.tbl.itemSelectionChanged.connect(self.fill_form)
        self.tbl.horizontalHeader().sectionClicked.connect(self.sort_table)

        self.refresh()

    def sort_table(self, column):
        self.tbl.sortItems(column, self.sort_order)
        self.sort_order = QtCore.Qt.DescendingOrder if self.sort_order == QtCore.Qt.AscendingOrder else QtCore.Qt.AscendingOrder

    def current_row_no(self):
        rows = self.tbl.selectionModel().selectedRows()
        if not rows: return None
        r = rows[0].row()
        item = self.tbl.item(r,0)
        return item.text() if item else None

    def fill_form(self):
        no = self.current_row_no()
        if no is None:
            self.clear_form()
            return
        with db_conn() as conn:
            c = conn.cursor()
            c.execute("""SELECT name,surname,class,branch,no,gender,phone,register_date FROM members WHERE no=?""", (no,))
            row = c.fetchone()
            if not row: return
            name,surname,klass,branch,no,gender,phone,regdate = row
            self.edName.setText(name or "")
            self.edSurname.setText(surname or "")
            self.edClass.setText(klass or "")
            self.edBranch.setText(branch or "")
            self.edNo.setText(no or "")
            self.cbGender.setCurrentText(gender or "")
            self.edPhone.setText(phone or "")
            if regdate:
                d = QDate.fromString(regdate, "yyyy-MM-dd")
                if d.isValid(): self.deRegister.setDate(d)
        
        self.btnUpdate.setDisabled(False)
        self.btnAdd.setDisabled(True)

    def clear_form(self):
        for w in [self.edName,self.edSurname,self.edClass,self.edBranch,self.edNo,self.edPhone,self.edSearch]:
            w.clear()
        self.cbGender.setCurrentIndex(0)
        self.deRegister.setDate(QDate.currentDate())
        self.tbl.clearSelection()
        self.btnUpdate.setDisabled(True)
        self.btnAdd.setDisabled(False)

    def on_add(self):
        name,surname,no = self.edName.text().strip(), self.edSurname.text().strip(), self.edNo.text().strip()
        if not (name and surname and no):
            QMessageBox.warning(self,"Uyarı","Ad, Soyad ve Numara zorunludur."); return
        data = (name,surname,self.edClass.text().strip(),self.edBranch.text().strip(),no,
                self.cbGender.currentText(),self.edPhone.text().strip(),
                self.deRegister.date().toString("yyyy-MM-dd"))
        try:
            with db_conn() as conn:
                c=conn.cursor()
                c.execute("""INSERT INTO members(name,surname,class,branch,no,gender,phone,register_date)
                             VALUES(?,?,?,?,?,?,?,?)""",data)
                conn.commit()
            self.refresh(); self.clear_form()
        except sqlite3.IntegrityError:
            QMessageBox.warning(self,"Hata","Bu numara zaten kayıtlı.")

    def on_update(self):
        no = self.current_row_no()
        if no is None: QMessageBox.information(self,"Bilgi","Güncellemek için satır seçin."); return
        data = (self.edName.text().strip(),self.edSurname.text().strip(),
                self.edClass.text().strip(),self.edBranch.text().strip(),
                self.cbGender.currentText(),self.edPhone.text().strip(),
                self.deRegister.date().toString("yyyy-MM-dd"), no)
        with db_conn() as conn:
            c=conn.cursor()
            c.execute("""UPDATE members SET name=?,surname=?,class=?,branch=?,gender=?,phone=?,register_date=? WHERE no=?""",data)
            conn.commit()
        self.refresh()

    def on_delete(self):
        selected_rows = self.tbl.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.information(self, "Bilgi", "Lütfen silmek için bir veya daha fazla üye seçin.")
            return

        reply = QMessageBox.question(self, "Silme Onayı",
                                     f"Seçili {len(selected_rows)} üyeyi silmek istediğinizden emin misiniz?",
                                     QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            nos_to_delete = [self.tbl.item(r.row(), 0).text() for r in selected_rows]
            with db_conn() as conn:
                c = conn.cursor()
                c.execute("DELETE FROM members WHERE no IN ({})".format(','.join('?' for _ in nos_to_delete)), nos_to_delete)
                conn.commit()
            self.refresh()
            self.clear_form()
            QMessageBox.information(self, "Başarılı", f"{len(nos_to_delete)} üye başarıyla silindi.")

    def on_delete_all_members(self):
        if self.user_role != 'admin':
            QMessageBox.warning(self, "Yetkisiz İşlem", "Bu işlemi yapmak için yönetici yetkisi gereklidir.")
            return

        reply = QMessageBox.question(self, "Tüm Üyeleri Sil",
                                    "**DİKKAT!** Bu işlem tüm üyeleri kalıcı olarak silecektir. Emin misiniz?",
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            with db_conn() as conn:
                c = conn.cursor()
                c.execute("DELETE FROM members")
                conn.commit()
            self.refresh()
            self.clear_form()
            QMessageBox.information(self, "Tamamlandı", "Tüm üyeler başarıyla silindi.")

    def export_excel(self):
        path,_=QFileDialog.getSaveFileName(self,"Excel'e aktar",os.path.join(EXPORT_DIR,"uyeler.xlsx"),"Excel (*.xlsx)")
        if not path: return
        headers=["AD","SOYAD","SINIF","ŞUBE","NUMARA","CİNSİYET","TELEFON","KAYIT TARİHİ"]
        with db_conn() as conn:
            c=conn.cursor()
            c.execute("SELECT name,surname,class,branch,no,gender,phone,register_date FROM members")
            rows=c.fetchall()
        wb=Workbook(); ws=wb.active; ws.append(headers)
        for r in rows: ws.append(list(r))
        wb.save(path); QMessageBox.information(self,"Tamam","Excel dosyası oluşturuldu.")

    def import_excel(self):
        path,_=QFileDialog.getOpenFileName(self,"Excel'den al",IMPORT_DIR,"Excel (*.xlsx)")
        if not path: return
        wb=load_workbook(path); ws=wb.active
        headers=[str(c.value).strip().upper() if c.value is not None else "" for c in ws[1]]
        expected=["AD","SOYAD","SINIF","ŞUBE","NUMARA","CİNSİYET","TELEFON","KAYIT TARİHİ"]
        missing=[h for h in expected if h not in headers]
        if missing:
            QMessageBox.warning(self,"Hata","Excel başlıkları uygun değil: " + ", ".join(missing)); return
        idx={h:headers.index(h) for h in expected}
        count=0
        with db_conn() as conn:
            c=conn.cursor()
            for row in ws.iter_rows(min_row=2,values_only=True):
                name=(row[idx["AD"]] or "").strip() if row[idx["AD"]] else ""
                surname=(row[idx["SOYAD"]] or "").strip() if row[idx["SOYAD"]] else ""
                no=(str(row[idx["NUMARA"]]).strip() if row[idx["NUMARA"]] is not None else "")
                if not (name and surname and no): continue
                klass = (row[idx["SINIF"]] or "") if row[idx["SINIF"]] else ""
                branch = (row[idx["ŞUBE"]] or "") if row[idx["ŞUBE"]] else ""
                gender = (row[idx["CİNSİYET"]] or "") if row[idx["CİNSİYET"]] else ""
                phone = (str(row[idx["TELEFON"]]).strip() if row[idx["TELEFON"]] is not None else "")
                reg = excel_val_to_iso(row[idx["KAYIT TARİHİ"]]) or datetime.date.today().isoformat()
                try:
                    c.execute("""INSERT OR IGNORE INTO members(name,surname,class,branch,no,gender,phone,register_date)
                                 VALUES(?,?,?,?,?,?,?,?)""",(name,surname,klass,branch,no,gender,phone,reg))
                    count += c.rowcount
                except Exception as e:
                    print("Üye satır hata:", e)
            conn.commit()
        self.refresh(); QMessageBox.information(self,"Tamam",f"Excel'den {count} kayıt eklendi.")

    def refresh(self):
        q = normalize(self.edSearch.text())
        with db_conn() as conn:
            c = conn.cursor()
            if q:
                c.execute("""SELECT no, name, surname, class, branch, gender, phone, register_date
                             FROM members WHERE normalize(name) LIKE ? OR normalize(surname) LIKE ? OR normalize(no) LIKE ?
                             ORDER BY name, surname ASC""", (f"%{q}%", f"%{q}%", f"%{q}%"))
            else:
                c.execute("""SELECT no, name, surname, class, branch, gender, phone, register_date
                             FROM members ORDER BY name, surname ASC""")
            rows = c.fetchall()
        self.tbl.setRowCount(len(rows))
        for r, row in enumerate(rows):
            for c_idx, val in enumerate(row):
                item = QTableWidgetItem(str(val))
                self.tbl.setItem(r, c_idx, item)
        self.tbl.resizeColumnsToContents()

# -------------------- Book Details Dialog --------------------
class BookDetailsDialog(QDialog):
    def __init__(self, book_id, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Kitap Detayları")
        self.setFixedSize(400, 300)
        
        layout = QFormLayout(self)
        self.lblTitle = QLabel()
        self.lblAuthor = QLabel()
        self.lblPublisher = QLabel()
        self.lblYear = QLabel()
        self.lblPages = QLabel()
        self.lblCategory = QLabel()
        self.lblBarcode = QLabel()
        self.lblDemirbas = QLabel()
        self.lblRaf = QLabel()
        self.lblDolap = QLabel()
        self.lblAdet = QLabel()
        self.lblNote = QLabel()
        
        layout.addRow("Kitap Adı:", self.lblTitle)
        layout.addRow("Yazar:", self.lblAuthor)
        layout.addRow("Yayınevi:", self.lblPublisher)
        layout.addRow("Basım Yılı:", self.lblYear)
        layout.addRow("Sayfa Sayısı:", self.lblPages)
        layout.addRow("Türü:", self.lblCategory)
        layout.addRow("Barkod:", self.lblBarcode)
        layout.addRow("Demirbaş No:", self.lblDemirbas)
        layout.addRow("Raf:", self.lblRaf)
        layout.addRow("Dolap:", self.lblDolap)
        layout.addRow("Mevcut Adet:", self.lblAdet)
        layout.addRow("Açıklama:", self.lblNote)
        
        self.load_details(book_id)

    def load_details(self, book_id):
        with db_conn() as conn:
            c = conn.cursor()
            c.execute("""SELECT title, author, publisher, year, pages, category, barcode, demirbas, raf, dolap, adet, note
                         FROM books WHERE id=?""", (book_id,))
            row = c.fetchone()
            if row:
                self.lblTitle.setText(row[0] or "Yok")
                self.lblAuthor.setText(row[1] or "Yok")
                self.lblPublisher.setText(row[2] or "Yok")
                self.lblYear.setText(str(row[3]) or "Yok")
                self.lblPages.setText(str(row[4]) or "Yok")
                self.lblCategory.setText(row[5] or "Yok")
                self.lblBarcode.setText(row[6] or "Yok")
                self.lblDemirbas.setText(row[7] or "Yok")
                self.lblRaf.setText(row[8] or "Yok")
                self.lblDolap.setText(row[9] or "Yok")
                self.lblAdet.setText(str(row[10]))
                self.lblNote.setText(row[11] or "Yok")
# -------------------------------------------------------------------------

# -------------------- Member History Widget --------------------
class MemberHistoryWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        layout = QVBoxLayout(self)
        
        self.tblHistory = QTableWidget(0, 4)
        self.tblHistory.setHorizontalHeaderLabels(["Kitap Adı", "Veriliş", "Son Tarih", "Teslim"])
        self.tblHistory.setSelectionBehavior(QTableWidget.SelectRows)
        self.tblHistory.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tblHistory.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        
        layout.addWidget(self.tblHistory)
        
    def load_history(self, member_id):
        self.tblHistory.setRowCount(0)
        if member_id is None:
            return
            
        with db_conn() as conn:
            c = conn.cursor()
            c.execute("""SELECT b.title, l.loan_date, l.due_date, l.return_date
                         FROM loans l
                         JOIN books b ON l.book_id = b.id
                         WHERE l.member_id = ?
                         ORDER BY l.return_date DESC, l.loan_date DESC""", (member_id,))
            rows = c.fetchall()
            
        self.tblHistory.setRowCount(len(rows))
        for r, row in enumerate(rows):
            for c_idx, val in enumerate(row):
                item = QTableWidgetItem(str(val) if val else "AKTİF")
                if val is None:
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)
                self.tblHistory.setItem(r, c_idx, item)
        self.tblHistory.resizeColumnsToContents()

# -------------------------------------------------------------------------


# -------------------- Loans Tab with preserved selection --------------------
class LoansTab(QWidget):
    def __init__(self, user_role, parent=None):
        super().__init__(parent)
        self.user_role = user_role
        self.sel_member_id = None
        self.sel_book_id = None
        self.selected_member_no = None
        self.selected_book_title = None
        self.active_rows = []
        self.build_ui()
        self.reset_loan_form() # Formu başlangıçta varsayılan tarihe ayarla

    def build_ui(self):
        main = QVBoxLayout(self)
        
        form_box = QGroupBox("Ödünç Ver")
        form = QFormLayout(form_box)
        
        # --- Sol ve Sağ Panelleri ayır ---
        form_panel_layout = QHBoxLayout()
        left_panel = QWidget()
        left_panel_layout = QFormLayout(left_panel)
        right_panel = QWidget()
        right_panel_layout = QVBoxLayout(right_panel)

        # Inputs (Left Panel)
        self.edMemberNo = QLineEdit(); self.edMemberNo.setPlaceholderText("Numara veya Ad Soyad Gir")
        self.member_suggest = QListWidget(); self.member_suggest.setFixedHeight(100); self.member_suggest.hide()
        self.edBookTitle = QLineEdit(); self.edBookTitle.setPlaceholderText("Kitap adı (yazdıkça ara ve seç)")
        self.book_suggest = QListWidget(); self.book_suggest.setFixedHeight(120); self.book_suggest.hide()
        
        book_input_layout = QHBoxLayout()
        book_input_layout.addWidget(self.edBookTitle)
        self.btnShowBookDetails = QPushButton("Detayları Göster")
        self.btnShowBookDetails.setDisabled(True)
        style_secondary(self.btnShowBookDetails)
        book_input_layout.addWidget(self.btnShowBookDetails)
        
        self.deLoan = QDateEdit(); self.deLoan.setCalendarPopup(True); self.deLoan.setDate(QDate.currentDate())
        self.deDue = QDateEdit(); self.deDue.setCalendarPopup(True); # Süre ayarından okunacak

        self.btnLoan = QPushButton("Ödünç Ver")
        style_primary(self.btnLoan)

        left_panel_layout.addRow("Üye No:", self.edMemberNo)
        left_panel_layout.addRow("", self.member_suggest)
        left_panel_layout.addRow("Kitap Adı:", book_input_layout)
        left_panel_layout.addRow("", self.book_suggest)
        left_panel_layout.addRow("Veriliş Tarihi:", self.deLoan)
        left_panel_layout.addRow("Son İade Tarihi:", self.deDue)
        left_panel_layout.addRow(self.btnLoan)

        # Member History (Right Panel)
        member_hist_box = QGroupBox("Seçili Üyenin Kitap Geçmişi")
        self.member_history_widget = MemberHistoryWidget()
        member_hist_layout = QVBoxLayout(member_hist_box)
        member_hist_layout.addWidget(self.member_history_widget)
        right_panel_layout.addWidget(member_hist_box)

        form_panel_layout.addWidget(left_panel)
        form_panel_layout.addWidget(right_panel)
        
        form.addRow(form_panel_layout)
        main.addWidget(form_box)

        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("Ödünç Arama:"))
        self.edActiveSearch = QLineEdit()
        self.edActiveSearch.setPlaceholderText("Üye, kitap, barkod vb. ara")
        search_layout.addWidget(self.edActiveSearch)
        search_layout.addStretch()
        main.addLayout(search_layout)

        # Tables
        self.tblActive = QTableWidget(0, 11)
        self.tblActive.setHorizontalHeaderLabels(["LoanID","Üye No","Üye","Sınıf","Şube","Kitap","Barkod","Raf","Dolap","Veriliş","Son Tarih"])
        self.tblActive.setSelectionBehavior(QTableWidget.SelectRows)
        self.tblActive.setEditTriggers(QTableWidget.NoEditTriggers)
        self.btnReturn = QPushButton("Seçiliyi Teslim Al")
        style_primary(self.btnReturn)

        hist_box = QGroupBox("Teslim Edilenler (Son 200)")
        self.tblHist = QTableWidget(0, 9)
        self.tblHist.setHorizontalHeaderLabels(["LoanID","Üye No","Üye","Sınıf","Şube","Kitap","Veriliş","Son Tarih","Teslim"])
        self.tblHist.setSelectionBehavior(QTableWidget.SelectRows)
        self.tblHist.setEditTriggers(QTableWidget.NoEditTriggers)

        main.addWidget(self.tblActive)
        main.addWidget(self.btnReturn)
        main.addWidget(hist_box)
        hist_layout = QVBoxLayout(hist_box)
        hist_layout.addWidget(self.tblHist)

        # Signals
        self.btnLoan.clicked.connect(self.on_loan)
        self.btnReturn.clicked.connect(self.on_return)
        self.deLoan.dateChanged.connect(self.update_due_date)
        self.edMemberNo.textChanged.connect(self.search_member_suggest)
        self.member_suggest.itemClicked.connect(self.pick_member)
        self.edBookTitle.textChanged.connect(self.search_book_suggest)
        self.book_suggest.itemClicked.connect(self.pick_book)
        self.btnShowBookDetails.clicked.connect(self.on_show_book_details)
        self.edActiveSearch.textChanged.connect(self.filter_active_loans)

        self.refresh_tables()

    # YENİ EKLEME: Varsayılan ödünç verme süresini dinamik olarak ayarlar.
    def reset_loan_form(self):
        default_days = get_default_loan_days()
        self.deLoan.setDate(QDate.currentDate())
        self.deDue.setDate(QDate.currentDate().addDays(default_days))
        self.edMemberNo.clear()
        self.edBookTitle.clear()
        self.sel_member_id = None
        self.sel_book_id = None
        self.selected_member_no = None
        self.selected_book_title = None
        self.member_history_widget.load_history(None)
        self.btnShowBookDetails.setDisabled(True)

    # YENİ EKLEME: Veriliş tarihi değiştiğinde Son İade tarihini günceller.
    def update_due_date(self, date):
        default_days = get_default_loan_days()
        self.deDue.setDate(date.addDays(default_days))

    def filter_active_loans(self):
        self.apply_active_filter()

    def apply_active_filter(self):
        search_text = self.edActiveSearch.text().strip()
        if not getattr(self, "active_rows", None):
            self.populate_active_table([])
            return

        if not search_text:
            filtered = self.active_rows
        else:
            norm_search = normalize(search_text)
            filtered = []
            for row in self.active_rows:
                for col in row:
                    if norm_search in normalize(str(col)):
                        filtered.append(row)
                        break

        self.populate_active_table(filtered)

    def populate_active_table(self, rows):
        self.tblActive.setRowCount(len(rows))
        today = datetime.date.today().isoformat()
        for r, row in enumerate(rows):
            due_date = row[10]
            is_overdue = due_date < today

            for c_idx, val in enumerate(row):
                item = QTableWidgetItem(str(val))
                if is_overdue:
                    item.setForeground(QtGui.QColor("red"))
                self.tblActive.setItem(r, c_idx, item)
        self.tblActive.resizeColumnsToContents()

    # --- live search helpers ---
    def search_member_suggest(self):
        text = self.edMemberNo.text().strip()
        if self.selected_member_no and text == self.selected_member_no:
            return
        self.sel_member_id = None
        self.selected_member_no = None
        self.member_suggest.clear()
        self.member_history_widget.load_history(None) # Üye bilgisi silindiğinde geçmişi temizle
        if not text:
            self.member_suggest.hide(); return
        
        q = normalize(text)
        with db_conn() as conn:
            c = conn.cursor()
            c.execute("""SELECT id,no,name,surname,class,branch FROM members
                         WHERE normalize(no) LIKE ? OR normalize(name) LIKE ? OR normalize(surname) LIKE ?
                         ORDER BY no LIMIT 50""", (f"%{q}%",f"%{q}%",f"%{q}%"))
            rows = c.fetchall()
        for mid,no,name,surname,klass,branch in rows:
            item = QtWidgets.QListWidgetItem(f"{no} — {name} {surname}  ({klass}{branch})")
            item.setData(Qt.UserRole, mid)
            self.member_suggest.addItem(item)
        self.member_suggest.setVisible(self.member_suggest.count()>0)

    def pick_member(self, item):
        mid = item.data(Qt.UserRole)
        member_text = item.text()
        
        self.edMemberNo.textChanged.disconnect(self.search_member_suggest)
        self.edMemberNo.setText(member_text)
        self.edMemberNo.textChanged.connect(self.search_member_suggest)
        
        self.selected_member_no = member_text.split(" — ", 1)[0]
        self.sel_member_id = int(mid)
        self.member_suggest.hide()
        self.member_history_widget.load_history(self.sel_member_id)

    def search_book_suggest(self):
        text = self.edBookTitle.text().strip()
        if self.selected_book_title and text == self.selected_book_title:
            return
        self.sel_book_id = None
        self.selected_book_title = None
        self.btnShowBookDetails.setDisabled(True)
        self.book_suggest.clear()
        if not text:
            self.book_suggest.hide(); return
        
        q = normalize(text)
        with db_conn() as conn:
            c = conn.cursor()
            c.execute("""SELECT id,title,author,adet FROM books
                         WHERE normalize(title) LIKE ?
                         ORDER BY title LIMIT 80""", (f"%{q}%",))
            rows = c.fetchall()
        for bid,title,author,adet in rows:
            item = QtWidgets.QListWidgetItem(f"{title} — {author or ''}  [Adet:{adet}]")
            item.setData(Qt.UserRole, (bid,title))
            self.book_suggest.addItem(item)
        self.book_suggest.setVisible(self.book_suggest.count()>0)

    def pick_book(self, item):
        bid,title = item.data(Qt.UserRole)
        
        self.edBookTitle.textChanged.disconnect(self.search_book_suggest)
        self.edBookTitle.setText(title)
        self.edBookTitle.textChanged.connect(self.search_book_suggest)

        self.sel_book_id = int(bid)
        self.selected_book_title = title
        self.book_suggest.hide()
        self.btnShowBookDetails.setDisabled(False)

    def on_show_book_details(self):
        if self.sel_book_id:
            dialog = BookDetailsDialog(self.sel_book_id, self)
            dialog.exec_()
        else:
            QMessageBox.warning(self, "Uyarı", "Lütfen önce bir kitap seçin.")

    # --- loan/return functions ---
    def on_loan(self):
        member_id = self.sel_member_id
        book_id = self.sel_book_id
        loan_date = self.deLoan.date().toString("yyyy-MM-dd")
        due_date = self.deDue.date().toString("yyyy-MM-dd")

        if member_id is None:
            QMessageBox.warning(self, "Uyarı", "Lütfen bir üye seçin.")
            return
        if book_id is None:
            QMessageBox.warning(self, "Uyarı", "Lütfen bir kitap seçin.")
            return
        
        loan_limit = get_loan_limit()

        with db_conn() as conn:
            c = conn.cursor()
            
            # Kitabın mevcut adedini kontrol et
            c.execute("SELECT adet FROM books WHERE id=?", (book_id,))
            book_adet = c.fetchone()[0]
            if book_adet <= 0:
                QMessageBox.warning(self, "Uyarı", "Bu kitabın mevcut kopyası yok.")
                return

            # Üyenin üzerindeki kitap sayısını kontrol et
            c.execute("SELECT COUNT(*) FROM loans WHERE member_id=? AND return_date IS NULL", (member_id,))
            loan_count = c.fetchone()[0]
            if loan_count >= loan_limit:
                QMessageBox.warning(self, "Uyarı", f"Bu üye, en fazla {loan_limit} kitap ödünç alabilir.")
                return

            # Ödünç kaydını ekle
            c.execute("""INSERT INTO loans(book_id, member_id, loan_date, due_date, return_date)
                         VALUES(?,?,?,?,?)""", (book_id, member_id, loan_date, due_date, None))

            # Kitabın adedini azalt
            c.execute("UPDATE books SET adet = adet - 1 WHERE id=?", (book_id,))
            conn.commit()

        # GÜNCELLEME: Ödünç verme işleminden sonra formu sıfırla
        self.reset_loan_form()
        self.refresh_tables()
        QMessageBox.information(self, "Başarılı", "Kitap ödünç verildi.")

    def on_return(self):
        rows = self.tblActive.selectionModel().selectedRows()
        if not rows:
            QMessageBox.information(self, "Bilgi", "Teslim almak için bir ödünç kaydı seçin.")
            return
        
        selected_row = rows[0].row()
        loan_id = int(self.tblActive.item(selected_row, 0).text())
        member_name = self.tblActive.item(selected_row, 2).text()
        book_title = self.tblActive.item(selected_row, 5).text()

        confirm_text = (
            f"{member_name} adlı üyenin\n"
            f"\"{book_title}\" kitabını teslim almak istediğinizden emin misiniz?"
        )

        if QMessageBox.question(self, "Onay", confirm_text) == QMessageBox.Yes:
            with db_conn() as conn:
                c = conn.cursor()

                # İlgili kitabın ID'sini bul
                c.execute("SELECT book_id FROM loans WHERE id=?", (loan_id,))
                book_id = c.fetchone()[0]

                # Teslim tarihini güncelle
                return_date = datetime.date.today().isoformat()
                c.execute("UPDATE loans SET return_date=? WHERE id=?", (return_date, loan_id))

                # Kitabın adedini artır
                c.execute("UPDATE books SET adet = adet + 1 WHERE id=?", (book_id,))
                conn.commit()

            self.refresh_tables()
            self.edActiveSearch.clear()
            QMessageBox.information(self, "Başarılı", "Kitap teslim alındı.")

    def refresh_tables(self):
        self.refresh_active_loans()
        self.refresh_loan_history()

    def refresh_active_loans(self):
        with db_conn() as conn:
            c = conn.cursor()
            c.execute("""SELECT
                            l.id, m.no, m.name || ' ' || m.surname, m.class, m.branch,
                            b.title, b.barcode, b.raf, b.dolap, l.loan_date, l.due_date
                         FROM
                            loans l
                         JOIN
                            members m ON l.member_id = m.id
                         JOIN
                            books b ON l.book_id = b.id
                         WHERE
                            l.return_date IS NULL
                         ORDER BY
                            l.due_date""")
            rows = c.fetchall()

        self.active_rows = rows
        self.apply_active_filter()

    def refresh_loan_history(self):
        with db_conn() as conn:
            c = conn.cursor()
            c.execute("""SELECT
                            l.id, m.no, m.name || ' ' || m.surname, m.class, m.branch,
                            b.title, l.loan_date, l.due_date, l.return_date
                         FROM
                            loans l
                         JOIN
                            members m ON l.member_id = m.id
                         JOIN
                            books b ON l.book_id = b.id
                         WHERE
                            l.return_date IS NOT NULL
                         ORDER BY
                            l.return_date DESC
                         LIMIT 200""")
            rows = c.fetchall()
        
        self.tblHist.setRowCount(len(rows))
        for r, row in enumerate(rows):
            for c_idx, val in enumerate(row):
                item = QTableWidgetItem(str(val))
                self.tblHist.setItem(r, c_idx, item)
        self.tblHist.resizeColumnsToContents()

# -------------------- Reports Tab --------------------
class ReportsTab(QWidget):
    def __init__(self, user_role, parent=None):
        super().__init__(parent)
        self.user_role = user_role
        self.build_ui()

    def build_ui(self):
        main = QVBoxLayout(self)

        top_row = QHBoxLayout()
        self.stat_box = QGroupBox("Genel İstatistikler")
        self.stat_layout = QFormLayout(self.stat_box)
        self.lblTotalBooks = QLabel("Toplam Kitap: 0")
        self.lblTotalMembers = QLabel("Toplam Üye: 0")
        self.lblActiveLoans = QLabel("Ödünçte Kitap: 0")
        self.stat_layout.addRow(self.lblTotalBooks)
        self.stat_layout.addRow(self.lblTotalMembers)
        self.stat_layout.addRow(self.lblActiveLoans)
        top_row.addWidget(self.stat_box)
        top_row.addStretch()

        self.btnRefresh = QPushButton("Raporları Yenile")
        style_secondary(self.btnRefresh)
        self.btnRefresh.clicked.connect(self.refresh_reports)

        pdf_box = QGroupBox("PDF Raporları")
        pdf_layout = QFormLayout(pdf_box)
        
        self.cbReportType = QComboBox()
        self.cbReportType.addItem("Geciken Kitaplar", "overdue")
        self.cbReportType.addItem("En Çok Ödünç Alınan Kitaplar", "top_books")
        self.cbReportType.addItem("Üye Ödünç Alma Sayıları", "member_loan_counts")
        self.cbReportType.addItem("Sınıfa/Şubeye Göre Üye Listesi", "class_list")

        date_layout = QHBoxLayout()
        self.deStartDate = QDateEdit(); self.deStartDate.setCalendarPopup(True); self.deStartDate.setDate(QDate.currentDate().addMonths(-1))
        self.deEndDate = QDateEdit(); self.deEndDate.setCalendarPopup(True); self.deEndDate.setDate(QDate.currentDate())
        date_layout.addWidget(QLabel("Başlangıç:"))
        date_layout.addWidget(self.deStartDate)
        date_layout.addWidget(QLabel("Bitiş:"))
        date_layout.addWidget(self.deEndDate)
        
        self.btnExportPDF = QPushButton("PDF Oluştur")
        style_primary(self.btnExportPDF)
        self.btnExportPDF.clicked.connect(self.export_pdf)

        pdf_layout.addRow("Rapor Türü:", self.cbReportType)
        pdf_layout.addRow("Tarih Aralığı:", date_layout)
        pdf_layout.addRow(self.btnExportPDF)
        
        top_row.addWidget(self.btnRefresh)
        top_row.addWidget(pdf_box)
        
        main.addLayout(top_row)
        
        self.tblOverdue = QTableWidget(0, 6)
        self.tblOverdue.setHorizontalHeaderLabels(["Üye No", "Üye Adı", "Kitap Adı", "Veriliş", "Son Tarih", "Gecikme (gün)"])
        self.tblOverdue.setSelectionBehavior(QTableWidget.SelectRows)
        self.tblOverdue.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tblOverdue.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)

        self.tblMostBorrowed = QTableWidget(0, 3)
        self.tblMostBorrowed.setHorizontalHeaderLabels(["Kitap Adı", "Yazar", "Ödünç Sayısı"])
        self.tblMostBorrowed.setSelectionBehavior(QTableWidget.SelectRows)
        self.tblMostBorrowed.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tblMostBorrowed.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        
        self.overdue_box = QGroupBox("Geciken Kitaplar")
        overdue_layout = QVBoxLayout(self.overdue_box)
        overdue_layout.addWidget(self.tblOverdue)

        self.most_borrowed_box = QGroupBox("En Çok Ödünç Verilen 10 Kitap")
        most_borrowed_layout = QVBoxLayout(self.most_borrowed_box)
        most_borrowed_layout.addWidget(self.tblMostBorrowed)

        main.addWidget(self.overdue_box)
        main.addWidget(self.most_borrowed_box)

        self.refresh_reports()

    def export_pdf(self):
        report_type = self.cbReportType.currentData()
        start_date = self.deStartDate.date().toString("yyyy-MM-dd")
        end_date = self.deEndDate.date().toString("yyyy-MM-dd")

        path, _ = QFileDialog.getSaveFileName(self, "PDF'e Aktar", os.path.join(REPORT_DIR, f"{report_type}_raporu.pdf"), "PDF (*.pdf)")
        if not path: return

        c = pdfcanvas.Canvas(path, pagesize=A4)
        c.setTitle("Kütüphane Yönetim Sistemi Raporu")
        c.setAuthor("Kütüphane Yönetim Sistemi")
        c.setCreator("Python PyQt5 App")

        y_pos = A4[1] - cm
        c.setFont(DEFAULT_FONT, 16)
        c.drawString(cm, y_pos, "Kütüphane Yönetim Sistemi Raporu")
        y_pos -= cm * 0.5
        c.setFont(DEFAULT_FONT, 10)
        c.drawString(cm, y_pos, f"Rapor Tarihi: {datetime.date.today().strftime('%d-%m-%Y')}")
        y_pos -= cm

        def draw_table(canvas, start_y, title, headers, data, col_widths):
            nonlocal y_pos
            y_pos = start_y
            
            c.setFont(DEFAULT_FONT + "-Bold", 12)
            c.drawString(cm, y_pos, title)
            y_pos -= cm
            
            # Draw headers
            c.setFont(DEFAULT_FONT + "-Bold", 10)
            x_pos = cm
            for h, w in zip(headers, col_widths):
                c.drawString(x_pos, y_pos, h)
                x_pos += w
            y_pos -= cm * 0.5

            c.setFont(DEFAULT_FONT, 8)
            
            for row in data:
                if y_pos < cm:
                    c.showPage()
                    y_pos = A4[1] - cm
                    c.setFont(DEFAULT_FONT + "-Bold", 12)
                    c.drawString(cm, y_pos, title + " (Devam)")
                    y_pos -= cm
                    c.setFont(DEFAULT_FONT + "-Bold", 10)
                    x_pos = cm
                    for h, w in zip(headers, col_widths):
                        c.drawString(x_pos, y_pos, h)
                        x_pos += w
                    y_pos -= cm * 0.5
                    c.setFont(DEFAULT_FONT, 8)

                x_pos = cm
                for c_idx, val in enumerate(row):
                    c.drawString(x_pos, y_pos, str(val))
                    x_pos += col_widths[c_idx]
                y_pos -= cm * 0.5
            return y_pos

        if report_type == "overdue":
            with db_conn() as conn:
                c_db = conn.cursor()
                today = datetime.date.today().isoformat()
                c_db.execute("""SELECT
                                m.no, m.name || ' ' || m.surname, b.title, l.loan_date, l.due_date,
                                CAST((JULIANDAY(?) - JULIANDAY(l.due_date)) AS INTEGER)
                                FROM loans l
                                JOIN members m ON l.member_id = m.id
                                JOIN books b ON l.book_id = b.id
                                WHERE l.return_date IS NULL AND l.due_date < ?
                                ORDER BY l.due_date ASC""", (today, today))
                rows = c_db.fetchall()
            
            headers = ["Üye No", "Üye Adı", "Kitap Adı", "Veriliş", "Son Tarih", "Gecikme (gün)"]
            col_widths = [2.5 * cm, 4 * cm, 4 * cm, 2.5 * cm, 2.5 * cm, 2 * cm]
            y_pos = draw_table(c, y_pos, "Geciken Kitaplar Raporu", headers, rows, col_widths)

        elif report_type == "top_books":
            with db_conn() as conn:
                c_db = conn.cursor()
                c_db.execute("""SELECT
                                b.title, b.author, COUNT(l.id)
                                FROM loans l
                                JOIN books b ON l.book_id = b.id
                                WHERE l.loan_date BETWEEN ? AND ?
                                GROUP BY b.id
                                ORDER BY COUNT(l.id) DESC
                                LIMIT 20""", (start_date, end_date))
                rows = c_db.fetchall()
            
            headers = ["Kitap Adı", "Yazar", "Ödünç Sayısı"]
            col_widths = [6 * cm, 6 * cm, 3 * cm]
            y_pos = draw_table(c, y_pos, f"En Çok Ödünç Alınan Kitaplar ({start_date} - {end_date})", headers, rows, col_widths)

        elif report_type == "member_loan_counts":
            with db_conn() as conn:
                c_db = conn.cursor()
                c_db.execute("""SELECT
                                m.no, m.name || ' ' || m.surname, COUNT(l.id)
                                FROM loans l
                                JOIN members m ON l.member_id = m.id
                                WHERE l.loan_date BETWEEN ? AND ?
                                GROUP BY m.id
                                ORDER BY COUNT(l.id) DESC""", (start_date, end_date))
                rows = c_db.fetchall()
            
            headers = ["Üye No", "Üye Adı", "Ödünç Sayısı"]
            col_widths = [4 * cm, 8 * cm, 4 * cm]
            y_pos = draw_table(c, y_pos, f"Üye Ödünç Alma Sayıları ({start_date} - {end_date})", headers, rows, col_widths)
        
        elif report_type == "class_list":
            with db_conn() as conn:
                c_db = conn.cursor()
                c_db.execute("""SELECT
                                no, name, surname, class, branch
                                FROM members
                                ORDER BY class, branch, surname, name""")
                rows = c_db.fetchall()
            
            headers = ["Numara", "Ad", "Soyad", "Sınıf", "Şube"]
            col_widths = [3 * cm, 4 * cm, 4 * cm, 2.5 * cm, 2 * cm]
            y_pos = draw_table(c, y_pos, "Sınıfa/Şubeye Göre Üye Listesi", headers, rows, col_widths)


        c.save()
        QMessageBox.information(self, "Başarılı", f"PDF dosyası oluşturuldu:\n{path}")

    def refresh_reports(self):
        self.refresh_stats()
        self.refresh_overdue_books()
        self.refresh_most_borrowed()

    def refresh_stats(self):
        with db_conn() as conn:
            c = conn.cursor()
            total_books = c.execute("SELECT COUNT(*) FROM books").fetchone()[0]
            total_members = c.execute("SELECT COUNT(*) FROM members").fetchone()[0]
            active_loans = c.execute("SELECT COUNT(*) FROM loans WHERE return_date IS NULL").fetchone()[0]
        
        self.lblTotalBooks.setText(f"Toplam Kitap: {total_books}")
        self.lblTotalMembers.setText(f"Toplam Üye: {total_members}")
        self.lblActiveLoans.setText(f"Ödünçte Kitap: {active_loans}")

    def refresh_overdue_books(self):
        with db_conn() as conn:
            c = conn.cursor()
            today = datetime.date.today().isoformat()
            c.execute("""SELECT
                            m.no, m.name || ' ' || m.surname, b.title, l.loan_date, l.due_date,
                            CAST((JULIANDAY(?) - JULIANDAY(l.due_date)) AS INTEGER)
                         FROM
                            loans l
                         JOIN
                            members m ON l.member_id = m.id
                         JOIN
                            books b ON l.book_id = b.id
                         WHERE
                            l.return_date IS NULL AND l.due_date < ?
                         ORDER BY
                            l.due_date ASC""", (today, today))
            rows = c.fetchall()
        
        self.tblOverdue.setRowCount(len(rows))
        for r, row in enumerate(rows):
            for c_idx, val in enumerate(row):
                item = QTableWidgetItem(str(val))
                self.tblOverdue.setItem(r, c_idx, item)
        self.tblOverdue.resizeColumnsToContents()

    def refresh_most_borrowed(self):
        with db_conn() as conn:
            c = conn.cursor()
            c.execute("""SELECT
                            b.title, b.author, COUNT(l.id)
                         FROM
                            loans l
                         JOIN
                            books b ON l.book_id = b.id
                         GROUP BY
                            b.id
                         ORDER BY
                            COUNT(l.id) DESC
                         LIMIT 10""")
            rows = c.fetchall()
        
        self.tblMostBorrowed.setRowCount(len(rows))
        for r, row in enumerate(rows):
            for c_idx, val in enumerate(row):
                item = QTableWidgetItem(str(val))
                self.tblMostBorrowed.setItem(r, c_idx, item)
        self.tblMostBorrowed.resizeColumnsToContents()

# -------------------- Settings Tab --------------------
class SettingsTab(QWidget):
    def __init__(self, user_role, user_id, parent=None):
        super().__init__(parent)
        self.user_role = user_role
        self.user_id = user_id
        self.build_ui()
        self.refresh_user_table()
        self.load_settings()
        if self.user_role != 'admin':
            self.user_management_box.hide()
            self.settings_box.hide()
    
    def clear_user_form(self):
        self.edNewUser.clear()
        self.edNewPassword.clear()
        self.cbNewUserRole.setCurrentIndex(0)
        self.tblUsers.clearSelection()
        self.edNewUser.setReadOnly(False)
        self.btnAddUser.setDisabled(False)
        self.btnUpdateUser.setDisabled(True)
        self.btnDeleteUser.setDisabled(True)

    def fill_user_form(self):
        selected_rows = self.tblUsers.selectionModel().selectedRows()
        if not selected_rows:
            self.clear_user_form()
            return
        
        r = selected_rows[0].row()
        user_id = int(self.tblUsers.item(r, 0).text())
        username = self.tblUsers.item(r, 1).text()
        role = self.tblUsers.item(r, 2).text()
        
        self.edNewUser.setText(username)
        self.cbNewUserRole.setCurrentText(role)
        self.edNewPassword.clear()
        
        self.edNewUser.setReadOnly(True)
        self.btnAddUser.setDisabled(True)
        self.btnUpdateUser.setDisabled(False)
        self.btnDeleteUser.setDisabled(False)

    def build_ui(self):
        main = QVBoxLayout(self)
        
        backup_box = QGroupBox("Veritabanı Yedekleme")
        backup_layout = QVBoxLayout(backup_box)
        
        info_label = QLabel("Uygulama her açıldığında otomatik yedekleme yapılır. İsterseniz şimdi manuel olarak da yedek alabilirsiniz.")
        self.btnBackup = QPushButton("Şimdi Yedekle")
        style_primary(self.btnBackup)
        
        backup_layout.addWidget(info_label)
        backup_layout.addWidget(self.btnBackup)
        
        self.user_management_box = QGroupBox("Kullanıcı Yönetimi")
        user_layout = QVBoxLayout(self.user_management_box)

        add_user_form = QFormLayout()
        self.edNewUser = QLineEdit()
        self.edNewPassword = QLineEdit()
        self.edNewPassword.setEchoMode(QLineEdit.Password)
        self.cbNewUserRole = QComboBox()
        self.cbNewUserRole.addItem("Görevli", "staff")
        self.cbNewUserRole.addItem("Yönetici", "admin")
        
        btn_layout = QHBoxLayout()
        self.btnAddUser = QPushButton("Kullanıcı Ekle")
        style_primary(self.btnAddUser)
        self.btnUpdateUser = QPushButton("Seçiliyi Güncelle")
        style_secondary(self.btnUpdateUser)
        self.btnDeleteUser = QPushButton("Seçiliyi Sil")
        style_danger(self.btnDeleteUser)
        self.btnUpdateUser.setDisabled(True)
        self.btnDeleteUser.setDisabled(True)
        
        btn_layout.addWidget(self.btnAddUser)
        btn_layout.addWidget(self.btnUpdateUser)
        btn_layout.addWidget(self.btnDeleteUser)

        add_user_form.addRow("Kullanıcı Adı:", self.edNewUser)
        add_user_form.addRow("Şifre:", self.edNewPassword)
        add_user_form.addRow("Rol:", self.cbNewUserRole)
        
        self.tblUsers = QTableWidget(0, 3)
        self.tblUsers.setHorizontalHeaderLabels(["ID", "Kullanıcı Adı", "Rol"])
        self.tblUsers.setSelectionBehavior(QTableWidget.SelectRows)
        self.tblUsers.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tblUsers.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        
        user_layout.addLayout(add_user_form)
        user_layout.addLayout(btn_layout)
        user_layout.addWidget(self.tblUsers)

        self.settings_box = QGroupBox("Genel Ayarlar")
        settings_layout = QFormLayout(self.settings_box)

        self.spLoanLimit = QSpinBox()
        self.spLoanLimit.setRange(1, 100)
        
        # YENİ EKLEME: Ödünç verme süresi spinbox'ı
        self.spDefaultLoanDays = QSpinBox()
        self.spDefaultLoanDays.setRange(1, 365)
        
        self.btnSaveSettings = QPushButton("Ayarları Kaydet")
        style_primary(self.btnSaveSettings)
        settings_layout.addRow("Ödünç Alma Sınırı:", self.spLoanLimit)
        # YENİ EKLEME: Ödünç verme süresi ayar satırı
        settings_layout.addRow("Varsayılan Ödünç Süresi (gün):", self.spDefaultLoanDays)
        settings_layout.addRow(self.btnSaveSettings)


        main.addWidget(backup_box)
        main.addWidget(self.user_management_box)
        main.addWidget(self.settings_box)
        main.addStretch()

        self.btnBackup.clicked.connect(self.backup_now)
        self.btnAddUser.clicked.connect(self.add_new_user)
        self.btnUpdateUser.clicked.connect(self.update_selected_user)
        self.btnDeleteUser.clicked.connect(self.delete_selected_user)
        self.tblUsers.itemSelectionChanged.connect(self.fill_user_form)
        self.btnSaveSettings.clicked.connect(self.on_save_settings)

    def load_settings(self):
        limit = get_loan_limit()
        self.spLoanLimit.setValue(limit)
        # YENİ EKLEME: Varsayılan ödünç süresini yükle
        default_days = get_default_loan_days()
        self.spDefaultLoanDays.setValue(default_days)

    def on_save_settings(self):
        new_limit = self.spLoanLimit.value()
        new_loan_days = self.spDefaultLoanDays.value() # YENİ EKLEME
        try:
            with db_conn() as conn:
                c = conn.cursor()
                c.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", ('loan_limit', str(new_limit)))
                c.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", ('default_loan_days', str(new_loan_days))) # YENİ EKLEME
                conn.commit()
            QMessageBox.information(self, "Başarılı", "Ayarlar başarıyla güncellendi.")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Ayarlar kaydedilirken bir hata oluştu:\n{e}")

    def refresh_user_table(self):
        self.clear_user_form() # Formu temizle ve butonları sıfırla
        with db_conn() as conn:
            c = conn.cursor()
            c.execute("SELECT id, username, role FROM users")
            rows = c.fetchall()
        
        self.tblUsers.setRowCount(len(rows))
        for r_idx, row in enumerate(rows):
            for c_idx, val in enumerate(row):
                item = QTableWidgetItem(str(val))
                self.tblUsers.setItem(r_idx, c_idx, item)
        self.tblUsers.resizeColumnsToContents()

    def backup_now(self):
        try:
            ts = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            dst = os.path.join(BACKUP_DIR, f"{ts}_kutuphane.db")
            shutil.copy2(DB_PATH, dst)
            QMessageBox.information(self, "Başarılı", f"Yedekleme başarılı!\nDosya: {dst}")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Yedekleme sırasında bir hata oluştu:\n{e}")

    def add_new_user(self):
        username = self.edNewUser.text().strip()
        password = self.edNewPassword.text().strip()
        role = self.cbNewUserRole.currentData()
        
        if not username or not password:
            QMessageBox.warning(self, "Uyarı", "Kullanıcı adı ve şifre boş bırakılamaz.")
            return

        try:
            hashed_password = hash_password(password)
            with db_conn() as conn:
                c = conn.cursor()
                c.execute("INSERT INTO users (username, password, role) VALUES (?, ?, ?)", (username, hashed_password, role))
                conn.commit()
            QMessageBox.information(self, "Başarılı", f"'{username}' ({'Yönetici' if role == 'admin' else 'Görevli'}) kullanıcısı başarıyla eklendi.")
            self.refresh_user_table()
        except sqlite3.IntegrityError:
            QMessageBox.warning(self, "Hata", "Bu kullanıcı adı zaten mevcut.")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Kullanıcı eklenirken bir hata oluştu:\n{e}")

    def update_selected_user(self):
        selected_rows = self.tblUsers.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "Uyarı", "Lütfen güncellemek için bir kullanıcı seçin.")
            return
            
        r = selected_rows[0].row()
        user_id = int(self.tblUsers.item(r, 0).text())
        new_role = self.cbNewUserRole.currentData()
        new_password = self.edNewPassword.text().strip()

        if user_id == self.user_id and new_role != 'admin':
            QMessageBox.warning(self, "Uyarı", "Kendi yönetici rolünüzü değiştiremezsiniz.")
            self.fill_user_form() # Formu eski haline getir
            return
            
        try:
            with db_conn() as conn:
                c = conn.cursor()
                if new_password:
                    hashed_password = hash_password(new_password)
                    c.execute("UPDATE users SET password=?, role=? WHERE id=?", (hashed_password, new_role, user_id))
                else:
                    c.execute("UPDATE users SET role=? WHERE id=?", (new_role, user_id))
                conn.commit()
            QMessageBox.information(self, "Başarılı", "Kullanıcı bilgileri başarıyla güncellendi.")
            self.refresh_user_table()
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Kullanıcı güncellenirken bir hata oluştu:\n{e}")

    def delete_selected_user(self):
        selected_rows = self.tblUsers.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "Uyarı", "Lütfen silmek için bir kullanıcı seçin.")
            return

        r = selected_rows[0].row()
        user_id = int(self.tblUsers.item(r, 0).text())
        username = self.tblUsers.item(r, 1).text()

        if user_id == self.user_id:
            QMessageBox.warning(self, "Uyarı", "Kendi hesabınızı silemezsiniz.")
            return

        reply = QMessageBox.question(self, "Onay",
                                     f"'{username}' kullanıcısını silmek istediğinizden emin misiniz?",
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            try:
                with db_conn() as conn:
                    c = conn.cursor()
                    c.execute("DELETE FROM users WHERE id=?", (user_id,))
                    conn.commit()
                QMessageBox.information(self, "Başarılı", "Kullanıcı başarıyla silindi.")
                self.refresh_user_table()
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Kullanıcı silinirken bir hata oluştu:\n{e}")

# -------------------- Login Dialog --------------------
class LoginDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Giriş Yap")
        self.setFixedSize(300, 150)
        self.setModal(True)
        self.user_role = None
        self.user_id = None

        layout = QFormLayout(self)
        
        self.edUsername = QLineEdit()
        self.edPassword = QLineEdit()
        self.edPassword.setEchoMode(QLineEdit.Password)

        self.btnLogin = QPushButton("Giriş")
        style_primary(self.btnLogin)
        self.btnExit = QPushButton("Çıkış")
        style_secondary(self.btnExit)
        
        layout.addRow("Kullanıcı Adı:", self.edUsername)
        layout.addRow("Şifre:", self.edPassword)
        
        button_layout = QHBoxLayout()
        button_layout.addWidget(self.btnLogin)
        button_layout.addWidget(self.btnExit)
        layout.addRow(button_layout)

        self.btnLogin.clicked.connect(self.check_credentials)
        self.btnExit.clicked.connect(self.reject)

        self.edUsername.returnPressed.connect(self.edPassword.setFocus)
        self.edPassword.returnPressed.connect(self.check_credentials)
        self.edUsername.setFocus()

    def check_credentials(self):
        username = self.edUsername.text().strip()
        password = self.edPassword.text().strip()
        hashed_password = hash_password(password)

        with db_conn() as conn:
            c = conn.cursor()
            c.execute("SELECT id, password, role FROM users WHERE username = ?", (username,))
            result = c.fetchone()
        
        if result and result[1] == hashed_password:
            self.user_id = result[0]
            self.user_role = result[2]
            self.accept()
        else:
            QMessageBox.warning(self, "Hata", "Kullanıcı adı veya şifre hatalı.")
            self.edPassword.clear()
            self.edUsername.setFocus()

# -------------------- Main Window --------------------
class LibraryApp(QMainWindow):
    def __init__(self, user_role, user_id):
        super().__init__()
        self.user_role = user_role
        self.user_id = user_id
        self.setWindowTitle(f"Kütüphane Yönetim Sistemi - ({self.user_role.capitalize()})")
        self.resize(1200, 800)
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.tabs = QTabWidget()
        
        # Sekme sırası değiştirildi
        self.tabs.addTab(LoansTab(self.user_role), "Ödünç İşlemleri")
        self.tabs.addTab(BooksTab(self.user_role), "Kitaplar")
        self.tabs.addTab(MembersTab(self.user_role), "Üyeler")
        self.tabs.addTab(ReportsTab(self.user_role), "Raporlar")
        
        # GÜNCELLEME: Sadece yönetici rolündeki kullanıcılar için "Ayarlar" sekmesini ekle
        if self.user_role == 'admin':
            self.tabs.addTab(SettingsTab(self.user_role, self.user_id), "Ayarlar")
        
        layout = QVBoxLayout(self.central_widget)
        layout.addWidget(self.tabs)

if __name__ == "__main__":
    init_db()
    auto_backup()
    app = QApplication(sys.argv)
    
    # Giriş ekranı ekle
    login = LoginDialog()
    if login.exec_() == QDialog.Accepted:
        user_id = login.user_id
        user_role = login.user_role
        ex = LibraryApp(user_role, user_id)
        ex.show()
        sys.exit(app.exec_())
    else:
        sys.exit(0)